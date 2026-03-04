from flask import Blueprint, render_template, session, redirect, request, url_for, abort, jsonify
from sqlalchemy import func, insert
from flask_session import Session
from sqlalchemy import select
import os
import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook

from app.tables.people import Person
from app.tables.project_people import ProjectPerson
from app.tables.users import User
from app.tables.files import File
from app.src.database import db
from app.tables.files import File
from app.tables.projects import Project
from app.src.util_functions import get_projects_for_user
from app.tables.people import Person
from app.tables.person_reports import PersonReport

DashBP= Blueprint('dashboard', __name__)
TIMELINE_ALLOWED_EXTENSIONS = {"csv", "xlsx", "xls"}


def _column_looks_unnamed(col_name):
    key = _canonicalize_col_name(col_name)
    return key.startswith("unnamed") or key == ""


def _detect_header_row(raw_df: pd.DataFrame, max_scan_rows: int = 60):
    max_rows = min(max_scan_rows, len(raw_df.index))
    if max_rows <= 0:
        return 0

    start_terms = {"start_date", "start", "begin_date", "begin", "from", "date"}
    end_terms = {"end_date", "end", "finish_date", "finish", "to", "due_date"}
    title_terms = {"title", "task", "event", "milestone", "name", "summary", "description"}

    best_row = 0
    best_score = -1

    for row_idx in range(max_rows):
        row_values = raw_df.iloc[row_idx].tolist()
        canonical = []
        for v in row_values:
            if pd.isna(v):
                continue
            key = _canonicalize_col_name(v)
            if key:
                canonical.append(key)

        if not canonical:
            continue

        non_unnamed = sum(1 for k in canonical if not _column_looks_unnamed(k))
        if non_unnamed == 0:
            continue

        has_start = any(k in start_terms for k in canonical)
        has_end = any(k in end_terms for k in canonical)
        has_title = any(k in title_terms for k in canonical)

        score = non_unnamed
        # Strongly prioritize rows that look like timeline headers even when
        # other rows have many non-empty cells (for example week matrices).
        if has_start:
            score += 200
        if has_end:
            score += 120
        if has_title:
            score += 80

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def _read_timeline_file_with_pandas(file_path: str):
    ext = file_path.split(".")[-1].lower()
    if ext == "csv":
        df = pd.read_csv(file_path)
        if not any(not _column_looks_unnamed(c) for c in df.columns):
            raw_df = pd.read_csv(file_path, header=None)
            header_row = _detect_header_row(raw_df)
            df = pd.read_csv(file_path, header=header_row)
        return df
    if ext in ["xlsx", "xls"]:
        df = pd.read_excel(file_path)
        # Many schedule sheets have title/banner rows above the actual headers.
        # If the default header is mostly "Unnamed", try detecting a better row.
        unnamed_count = sum(1 for c in df.columns if _column_looks_unnamed(c))
        if unnamed_count >= max(1, int(len(df.columns) * 0.5)):
            raw_df = pd.read_excel(file_path, header=None)
            header_row = _detect_header_row(raw_df)
            df = pd.read_excel(file_path, header=header_row)
        return df
    raise ValueError("Unsupported file type.")


def _canonicalize_col_name(name: str):
    return re.sub(r"[^a-z0-9]+", "_", str(name).strip().lower()).strip("_")


def _to_datetime_series(series: pd.Series):
    # Parse textual date-like values first. Treat numeric values separately so
    # plain counters (1,2,3,...) are not interpreted as epoch nanoseconds.
    as_text = series.astype("string").str.strip()
    as_text = as_text.replace(
        {
            "": pd.NA,
            "nan": pd.NA,
            "nat": pd.NA,
            "none": pd.NA,
            "null": pd.NA,
            "n/a": pd.NA,
        }
    )
    try:
        dt = pd.to_datetime(as_text, errors="coerce", format="mixed")
    except TypeError:
        dt = pd.to_datetime(as_text, errors="coerce")

    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.notna().any():
        numeric_intlike = numeric.notna() & (numeric % 1 == 0)

        # Excel serial dates (roughly years 1954-2149).
        excel_mask = numeric.between(20000, 90000)
        if excel_mask.any():
            excel_dt = pd.to_datetime(numeric.where(excel_mask), unit="D", origin="1899-12-30", errors="coerce")
            dt = dt.where(~excel_mask, excel_dt)

        # YYYYMMDD integers.
        ymd_mask = numeric_intlike & numeric.between(19000101, 29991231)
        if ymd_mask.any():
            ymd_text = numeric.where(ymd_mask).astype("Int64").astype("string")
            ymd_dt = pd.to_datetime(ymd_text, format="%Y%m%d", errors="coerce")
            dt = dt.where(~ymd_mask, ymd_dt)

        # Unix epoch seconds and milliseconds.
        epoch_sec_mask = numeric_intlike & numeric.between(946684800, 4102444800)
        if epoch_sec_mask.any():
            sec_dt = pd.to_datetime(numeric.where(epoch_sec_mask), unit="s", errors="coerce")
            dt = dt.where(~epoch_sec_mask, sec_dt)

        epoch_ms_mask = numeric_intlike & numeric.between(946684800000, 4102444800000)
        if epoch_ms_mask.any():
            ms_dt = pd.to_datetime(numeric.where(epoch_ms_mask), unit="ms", errors="coerce")
            dt = dt.where(~epoch_ms_mask, ms_dt)

    return dt


def _coerce_week_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        if float(value).is_integer():
            wk = int(value)
            if wk > 0:
                return wk
        return None

    text = str(value).strip().lower()
    if not text:
        return None

    week_match = re.fullmatch(r"week\s*(\d+)", text)
    if week_match:
        return int(week_match.group(1))

    numeric_match = re.fullmatch(r"\d+(?:\.0+)?", text)
    if numeric_match:
        wk = int(float(text))
        return wk if wk > 0 else None

    return None


def _cell_has_timeline_marker(cell):
    value = cell.value
    if value is not None and str(value).strip() != "":
        return True

    fill = getattr(cell, "fill", None)
    if not fill:
        return False

    return fill.patternType not in (None, "none")


def _find_week_matrix_header(ws, max_scan_rows: int = 120, max_scan_cols: int = 80):
    scan_rows = min(max_scan_rows, ws.max_row)
    scan_cols = min(max_scan_cols, ws.max_column)

    for row_idx in range(1, scan_rows + 1):
        for col_idx in range(1, scan_cols + 1):
            cell_value = ws.cell(row_idx, col_idx).value
            if _canonicalize_col_name(cell_value) != "group_member_assigned":
                continue

            week_cols = []
            for candidate_col in range(col_idx + 1, scan_cols + 1):
                wk = _coerce_week_number(ws.cell(row_idx, candidate_col).value)
                if wk is None:
                    if week_cols:
                        break
                    continue
                week_cols.append((candidate_col, wk))

            if len(week_cols) >= 2:
                return row_idx, col_idx, week_cols

    return None, None, []


def _build_week_matrix_timeline(file_path: str):
    ext = file_path.split(".")[-1].lower()
    if ext not in ["xlsx", "xls"]:
        raise ValueError("Week-matrix timeline fallback supports Excel files only.")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    header_row, owner_col, week_cols = _find_week_matrix_header(ws)
    if not header_row or not owner_col or not week_cols:
        raise ValueError("Could not locate a week-matrix timeline layout.")

    task_col = max(1, owner_col - 1)
    # Week 1 anchored to Monday of the current year so week spans are calendar-aligned.
    jan1 = datetime(datetime.utcnow().year, 1, 1)
    week_one = jan1 - timedelta(days=jan1.weekday())

    events = []
    task_rows_seen = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_task = ws.cell(row_idx, task_col).value
        if raw_task is None:
            continue

        task = str(raw_task).strip()
        if not task:
            continue

        canonical_task = _canonicalize_col_name(task)
        if canonical_task in {"key", "week", "group_member_assigned"}:
            continue
        if task.startswith("["):
            # Section headers in this template are bracketed and should not be events.
            continue

        task_rows_seen += 1
        active_weeks = []
        for col_idx, week_number in week_cols:
            if _cell_has_timeline_marker(ws.cell(row_idx, col_idx)):
                active_weeks.append(week_number)

        if not active_weeks:
            continue

        owner = ws.cell(row_idx, owner_col).value
        owner_text = str(owner).strip() if owner is not None else ""
        content = task if not owner_text else f"{task} ({owner_text})"

        start_week = min(active_weeks)
        end_week = max(active_weeks)
        start_dt = week_one + timedelta(weeks=start_week - 1)
        end_dt = week_one + timedelta(weeks=end_week)

        events.append(
            {
                "id": len(events) + 1,
                "content": content,
                "start": start_dt.isoformat(),
                "end": end_dt.isoformat(),
                "missing_start": False,
            }
        )

    if not events:
        raise ValueError("No timeline events found in week matrix.")

    meta = {
        "rows_total": int(task_rows_seen),
        "events_total": int(len(events)),
        "title_col": f"Column {task_col}",
        "start_col": "Week columns (colored cells)",
        "end_col": "Week columns (colored cells)",
        "parse_mode": "week_matrix",
    }
    return events, meta


def _date_parse_stats(series: pd.Series, ignore_first_row: bool = False):
    series_to_check = series.iloc[1:] if ignore_first_row and len(series.index) > 1 else series
    non_null_count = int(series_to_check.notna().sum())
    if non_null_count == 0:
        return 0, 0, 0.0

    parsed = _to_datetime_series(series_to_check)
    parsed_count = int(parsed.notna().sum())
    ratio = parsed_count / non_null_count
    return non_null_count, parsed_count, ratio


def _infer_timeline_columns(df: pd.DataFrame):
    if df.empty:
        raise ValueError("The uploaded file is empty.")

    normalized = {_canonicalize_col_name(c): c for c in df.columns}

    def _pick_exact(preferred: list[str]):
        for candidate in preferred:
            key = _canonicalize_col_name(candidate)
            if key in normalized:
                return normalized[key]
        return None

    def _pick_contains(required_terms: list[str]):
        for key, original in normalized.items():
            if all(term in key for term in required_terms):
                return original
        return None

    inference_df = df.iloc[1:] if len(df.index) > 1 else df

    def _has_any_parseable_date(col_name: str | None):
        if not col_name:
            return False
        _, parsed_count, _ = _date_parse_stats(df[col_name], ignore_first_row=True)
        return parsed_count >= 1

    def _is_reliable_date_col(col_name: str | None):
        if not col_name:
            return False
        non_null_count, parsed_count, ratio = _date_parse_stats(df[col_name], ignore_first_row=True)
        # Require multiple parseable values for confidence.
        return non_null_count >= 2 and parsed_count >= 2 and ratio >= 0.25

    # Primary rule: assume 1st=task, 2nd=start, 3rd=end.
    populated_cols = [col for col in df.columns if int(inference_df[col].notna().sum()) > 0]
    ordered_cols = populated_cols if populated_cols else list(df.columns)

    title_col = ordered_cols[0] if len(ordered_cols) >= 1 else None
    start_col = ordered_cols[1] if len(ordered_cols) >= 2 else None
    end_col = ordered_cols[2] if len(ordered_cols) >= 3 else None

    if start_col and not _has_any_parseable_date(start_col):
        start_col = None
    if end_col and not _has_any_parseable_date(end_col):
        end_col = None

    # Secondary rule: infer by semantic headers.
    semantic_start_col = _pick_exact(["start_date", "start", "begin_date", "begin", "from", "date"])
    semantic_end_col = _pick_exact(["end_date", "end", "finish_date", "finish", "to", "due_date"])
    semantic_title_col = _pick_exact(["title", "task", "event", "milestone", "name", "summary", "description"])

    if not semantic_start_col:
        semantic_start_col = _pick_contains(["start", "date"]) or _pick_contains(["start"])
    if not semantic_end_col:
        semantic_end_col = _pick_contains(["end", "date"]) or _pick_contains(["finish"])
    if not semantic_title_col:
        semantic_title_col = _pick_contains(["task"]) or _pick_contains(["event"]) or _pick_contains(["name"])

    if not start_col and semantic_start_col and _is_reliable_date_col(semantic_start_col):
        start_col = semantic_start_col
    if not end_col and semantic_end_col and _is_reliable_date_col(semantic_end_col):
        end_col = semantic_end_col
    if not title_col and semantic_title_col:
        title_col = semantic_title_col

    if not start_col:
        best_candidate = None
        for col in df.columns:
            non_null_count, parsed_count, ratio = _date_parse_stats(df[col], ignore_first_row=True)
            if non_null_count < 2 or parsed_count < 2:
                continue
            if ratio < 0.25:
                continue
            score = (parsed_count, ratio)
            if best_candidate is None or score > best_candidate[0]:
                best_candidate = (score, col)

        if best_candidate:
            start_col = best_candidate[1]

    if not end_col and start_col:
        best_candidate = None
        for col in df.columns:
            if col == start_col:
                continue
            non_null_count, parsed_count, ratio = _date_parse_stats(df[col], ignore_first_row=True)
            if non_null_count < 2 or parsed_count < 2:
                continue
            if ratio < 0.25:
                continue
            score = (parsed_count, ratio)
            if best_candidate is None or score > best_candidate[0]:
                best_candidate = (score, col)

        if best_candidate:
            end_col = best_candidate[1]

    if not title_col:
        for col in df.columns:
            if col not in [start_col, end_col]:
                title_col = col
                break

    if not start_col:
        preview_cols = ", ".join(str(c) for c in list(df.columns)[:12])
        raise ValueError(
            "Could not infer a start-date column. Expected col2=start date by default, or a named column like start_date/date. "
            f"Detected columns: {preview_cols}"
        )

    return title_col, start_col, end_col


def _build_timeline_events(df: pd.DataFrame, title_col: str | None, start_col: str, end_col: str | None):
    starts = _to_datetime_series(df[start_col])
    ends = _to_datetime_series(df[end_col]) if end_col else None

    if title_col:
        titles = df[title_col].fillna("").astype(str).str.strip()
    else:
        titles = pd.Series([""] * len(df))

    events = []
    for idx in range(len(df)):
        start_ts = starts.iloc[idx]
        start_missing = pd.isna(start_ts)
        if pd.isna(start_ts):
            if ends is not None and not pd.isna(ends.iloc[idx]):
                start_ts = ends.iloc[idx]
            else:
                continue

        title = titles.iloc[idx] if idx < len(titles) else ""
        if not title:
            title = f"Event {idx + 1}"

        end_iso = None
        if ends is not None:
            end_ts = ends.iloc[idx]
            if not pd.isna(end_ts):
                if end_ts < start_ts:
                    end_ts = start_ts
                if end_ts != start_ts:
                    end_iso = end_ts.isoformat()

        events.append(
            {
                "id": idx + 1,
                "content": title,
                "start": start_ts.isoformat(),
                "end": end_iso,
                "missing_start": bool(start_missing),
            }
        )

    if not events:
        raise ValueError("No valid timeline events found. Check your date columns.")

    return events


def _file_extension(file_name: str):
    if not file_name or "." not in file_name:
        return ""
    return file_name.rsplit(".", 1)[-1].lower()


def _get_project_timeline_files(project_id: str):
    project_files = (
        db.session.execute(
            select(File)
            .where(File.project_id == project_id)
            .order_by(File.upload_date.desc())
        )
        .scalars()
        .all()
    )

    timeline_files = []
    for f in project_files:
        ext = _file_extension(f.file_name_original) or _file_extension(f.file_name_disk)
        if ext in TIMELINE_ALLOWED_EXTENSIONS:
            timeline_files.append(f)

    return timeline_files


def _resolve_timeline_file_path(file_row: File):
    configured_root = os.getenv("FILE_UPLOAD_STORAGE_PATH", "").strip()
    if not configured_root:
        raise ValueError("FILE_UPLOAD_STORAGE_PATH is not configured.")

    disk_path = os.path.join(configured_root, file_row.file_name_disk)
    if not os.path.exists(disk_path):
        raise ValueError(f"File not found on disk: {file_row.file_name_disk}")

    return disk_path

@DashBP.before_request
def require_login():
    if "user_id" not in session:
        return redirect(url_for("authentication.login", next=request.path))

@DashBP.route('/')
def get_dashboard_main():
    user_id = session["user_id"]

    user = db.session.get(User, user_id)
    username = user.username if user else "User"

    projects = get_projects_for_user(user_id)

    dashboard_title = f'Welcome, {username}'
    description = "Here are your projects."

    return render_template(
        "dashboard/dashboard_overview.html",
        dashboard_title=dashboard_title,
        description=description,
        projects=projects,
    ), 200

@DashBP.route('/<project_id>/')
def get_dashboard_project_home(project_id):
    project = db.session.get(Project, project_id)
    if not project:
        return abort(404)

    people_rows = (
        db.session.execute(
            select(Person, ProjectPerson)
            .join(ProjectPerson, ProjectPerson.person_id == Person.id)
            .where(ProjectPerson.project_id == project_id)
            .order_by(Person.name.asc())
        )
        .all()
    )

    recent_files = (
        db.session.execute(
            select(File)
            .where(File.project_id == project_id)
            .order_by(File.upload_date.desc())
            .limit(6)
        )
        .scalars()
        .all()
    )

    return render_template(
        "dashboard/dashboard_project_home.html",
        project=project,
        active_project_id=project.id,
        people_rows=people_rows,
        recent_files=recent_files,
    ), 200

@DashBP.route('/<project_id>/visualizations/') 
def get_dashboard_project_visualizations(project_id):
    project = db.session.get(Project, project_id)

    if not project:
        return abort(404)

    return render_template(
        "dashboard/dashboard_visualizations.html",
        project=project,
        active_project_id=project.id
    ), 200


@DashBP.route('/<project_id>/timeline/', methods=['GET'])
def get_dashboard_project_timeline(project_id):
    project = db.session.get(Project, project_id)

    if not project:
        return abort(404)

    timeline_files = _get_project_timeline_files(project_id)
    selected_file_id = request.args.get("file_id", "").strip()
    selected_file = None

    timeline_events = []
    timeline_meta = None
    error = None
    success = None

    if timeline_files:
        files_by_id = {f.id: f for f in timeline_files}
        if selected_file_id:
            selected_file = files_by_id.get(selected_file_id)
            if not selected_file:
                error = "Selected file is not available for this project."
        else:
            selected_file = timeline_files[0]
            selected_file_id = selected_file.id
    else:
        error = "No timeline-compatible files found. Upload a CSV/XLS/XLSX file from the Files tab."

    if selected_file:
        try:
            disk_path = _resolve_timeline_file_path(selected_file)
            try:
                df = _read_timeline_file_with_pandas(disk_path)
                title_col, start_col, end_col = _infer_timeline_columns(df)
                timeline_events = _build_timeline_events(df, title_col, start_col, end_col)

                timeline_meta = {
                    "rows_total": int(len(df.index)),
                    "events_total": int(len(timeline_events)),
                    "title_col": title_col or "(auto-generated)",
                    "start_col": start_col,
                    "end_col": end_col or "(none)",
                    "file_path": disk_path,
                    "file_name_original": selected_file.file_name_original,
                    "file_uploaded_at": selected_file.upload_date.isoformat() if selected_file.upload_date else None,
                }
                success = f"Loaded timeline from {selected_file.file_name_original}."
            except Exception:
                # Fallback for matrix-style timeline sheets where weeks are represented
                # by colored cells under week-number columns.
                fallback_events, fallback_meta = _build_week_matrix_timeline(disk_path)
                timeline_events = fallback_events
                timeline_meta = {
                    "rows_total": fallback_meta["rows_total"],
                    "events_total": fallback_meta["events_total"],
                    "title_col": fallback_meta["title_col"],
                    "start_col": fallback_meta["start_col"],
                    "end_col": fallback_meta["end_col"],
                    "file_path": disk_path,
                    "file_name_original": selected_file.file_name_original,
                    "file_uploaded_at": selected_file.upload_date.isoformat() if selected_file.upload_date else None,
                }
                success = (
                    f"Loaded timeline from {selected_file.file_name_original} "
                    "(week-matrix format detected)."
                )
        except Exception as exc:
            error = f"Could not build timeline: {exc}"

    return render_template(
        "dashboard/dashboard_timeline.html",
        project=project,
        active_project_id=project.id,
        timeline_events=timeline_events,
        timeline_meta=timeline_meta,
        timeline_files=timeline_files,
        selected_file_id=selected_file_id,
        error=error,
        success=success,
    ), 200


@DashBP.route('/<project_id>/people/')
def get_dashboard_project_people(project_id):
    project = db.session.get(Project, project_id)

    if not project:
        return abort(404)

    xlsx_files = db.session.query(File).filter(
        File.project_id == project_id,
        func.lower(File.file_name_original).like('%.xlsx')
    ).order_by(File.upload_date.desc()).all()
    
    # reporting_people = [
    #     {"id": "person_1", "name": "Bryan Coblentz", "title": "Project Manager"},
    #     {"id": "person_2", "name": "Kevin Hare", "title": "Tech Lead"},
    #     {"id": "person_3", "name": "Matt Troyer", "title": "CEO"},
    #     {"id": "person_4", "name": "Jamie Coblentz", "title": "Software Engineer"},
    #     {"id": "person_5", "name": "Merl Coblentz", "title": "Software Engineer"},
    #     {"id": "person_6", "name": "Traci Miller", "title": "Software Engineer"},
    #     {"id": "person_7", "name": "Joel Coblentz", "title": "Product Designer"},
    #     {"id": "person_8", "name": "Joe Yoder", "title": "UX Desiner"},
    #     {"id": "person_9", "name": "Teresa Bonifant", "title": "Software Engineer"},
    #     {"id": "person_10", "name": "Darrin Hess", "title": "Consulting"},
    # ]


    project_member_ids = select(ProjectPerson.person_id).where(
        ProjectPerson.project_id == project_id
    )

    reporting_edges = (
        db.session.query(PersonReport.person_id, PersonReport.reports_to_id)
        .filter(PersonReport.person_id.in_(project_member_ids))
        .filter(PersonReport.reports_to_id.in_(project_member_ids))
        .all()
    )
    reporting_links = {f"{person_id}:{reports_to_id}" for person_id, reports_to_id in reporting_edges}

    people_rows = (
            db.session.execute(
                select(Person, ProjectPerson)
                .join(ProjectPerson, ProjectPerson.person_id == Person.id)
                .where(ProjectPerson.project_id == project_id)
                .order_by(Person.name.asc())
                )
            .all()
            )
    
    return render_template(
        "dashboard/dashboard_people.html",
        project=project,
        active_project_id=project.id,
        people_rows=people_rows,
        reporting_links=reporting_links
        ), 200

@DashBP.route('/<project_id>/people/updatematrix', methods=['POST'])
def update_reporting_matrix(project_id):
    payload = request.get_json()

    person_id = payload["person_id"]
    manager_id = payload["manager_id"]
    checked = payload["checked"]

    is_checked = str(checked).lower() == "true" if isinstance(checked, str) else bool(checked)

    if is_checked:
        db.session.execute(
            insert(PersonReport)
            .values(person_id=person_id, reports_to_id=manager_id)
        )
    else:
        db.session.query(PersonReport).filter(
            PersonReport.person_id == person_id,
            PersonReport.reports_to_id == manager_id
        ).delete()

    db.session.commit()

    return jsonify({
        "person_id": person_id,
        "manager_id": manager_id,
        "checked": is_checked
    }), 200
