from flask import Blueprint, render_template, session, redirect, request, url_for, jsonify
from sqlalchemy import select, exists, and_
from app.tables.people import Person
from app.tables.project_people import ProjectPerson
from openpyxl import load_workbook
import csv
from app.src.database import db
from app.tables.projects import Project
from app.tables.roles import Role
from app.tables.files import File
from app.src.util_functions import get_projects_for_user
from werkzeug.utils import secure_filename
import os
from app.src.constants import ALLOWED_FILE_EXTENSIONS
import uuid

ProjectsBP = Blueprint('projects', __name__)

# @projectsBP.before_request
# def require_login():
#     if "user_id" not in session:
#         return redirect(url_for("authentication.login", next=request.path))

# TEMPORARY test project data
# will be injected into database when you try to access /projects/createdummies endpoint
dummy_data = {
    'Sherwin Williams DBMS Transition': '''This project involves migration Sherwin Williams\' customer service database from MongoDB to PostreSQL. Hired on June 12, expect contract to be ~6 months. Emergency contact at Sherwin: Timothy Harris 330 111 1111''',
    'Huntington Bank ATM API': '''Spending a lot of time digging into how their ATM stack is actually wired together, which is messier than it looks on paper. Talking with ops about weird edge cases (timeouts, partial transactions, logging gaps) and sketching a cleaner flow for updates and testing so releases don’t feel like mini heart attacks. The main aim right now is just making the system more predictable and less painful to maintain.''',
    'Diebold Nixdorf AI Chatbot': 'Focused on how the chatbot should behave in real situations instead of the idealized flows. Sitting in on support calls, noting common customer frustrations, and translating that into conversation logic that doesn’t feel robotic. Coordinating with the vendor on integrations and agent handoff. It’s still iterative, but early pilots show a drop in repetitive tickets.'
}

def _user_has_project_access(user_id: str, project_id: str) -> bool:
    return db.session.execute(
        select(
            exists().where(
                and_(
                    Role.user_id == user_id,
                    Role.project_id == project_id,
                )
            )
        )
    ).scalar()


def _parse_csv_headers_preview(filepath: str, preview_rows: int = 10):
    with open(filepath, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = []
        for _, row in zip(range(preview_rows), reader):
            rows.append(row)
        return headers, rows


def _parse_xlsx_headers_preview(filepath: str, preview_rows: int = 10):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    headers = [("" if h is None else str(h)).strip() for h in (headers_raw or [])]

    rows = []
    for _, r in zip(range(preview_rows), rows_iter):
        rows.append([("" if v is None else str(v)).strip() for v in r])

    return headers, rows


def _read_all_csv_rows(filepath: str):
    with open(filepath, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = [row for row in reader]
        return headers, rows


def _read_all_xlsx_rows(filepath: str):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    headers = [("" if h is None else str(h)).strip() for h in (headers_raw or [])]

    rows = []
    for r in rows_iter:
        rows.append([("" if v is None else str(v)).strip() for v in r])

    return headers, rows

@ProjectsBP.route('/new', methods=['GET', 'POST'])
def create_project():
    user_id = session["user_id"]
    projects = get_projects_for_user(user_id)
    
    if request.method == 'GET':
        return render_template("pages/project_create.html", projects=projects, status=None), 200
    elif request.method == 'POST':
        title = request.form['title']
        description = request.form['description']

        proj_id = str(uuid.uuid4())
        project = Project(id=proj_id, owner_id=user_id,title=title, description=description)
        db.session.add(project)

        r = Role(user_id = user_id, project_id = proj_id, role = 'owner')
        db.session.add(r)

        db.session.commit()
        projects = get_projects_for_user(user_id)

        return render_template('dashboard/dashboard_overview.html', projects=projects, status='success'), 200

@ProjectsBP.route('/edit/<project_id>', methods=['GET', 'POST'])
def edit_project(project_id):
    if request.method == 'GET':
        project = db.session.get(Project, project_id)
        return render_template("pages/project_add_data.html", active_project_id=project.id, active_project=project, status=None), 200

    elif request.method == 'POST':
        which_form = request.form.get('which_form', 'update_fields')

        if which_form == 'update_fields':
            project = db.session.get(Project, project_id)
            project.title = request.form['title']
            project.description = request.form['description']
            db.session.commit()

            return render_template("dashboard/dashboard_overview.html", dashboard_title='Hello, Username!', projects=get_projects_for_user(user_id=session["user_id"])), 200

        return redirect(url_for('projects.edit_project', project_id=project_id))


@ProjectsBP.route('/<project_id>/files', methods=['GET'])
def project_files_list(project_id):
    files = db.session.query(File).filter(File.project_id == project_id).order_by(File.upload_date.desc()).all()
    project = db.session.get(Project, project_id)
    return render_template("dashboard/dashboard_project_files.html", project=project, active_project_id=project.id, files=files), 200


@ProjectsBP.route('/<project_id>/files/upload', methods=['GET', 'POST'])
def project_files_upload(project_id):
    project = db.session.get(Project, project_id)

    if request.method == 'GET':
        return render_template("pages/project_upload.html", active_project=project, active_project_id=project.id, status=None), 200

    f = request.files.get('uploaded_file')
    if not f or f.filename == '':
        return render_template("error/404.html"), 404

    ext = f.filename.split('.')[-1].lower()
    if ext not in ALLOWED_FILE_EXTENSIONS:
        return render_template("pages/project_upload.html", active_project=project, active_project_id=project.id, error="Invalid extension"), 400

    file_name = secure_filename(f.filename)

    temp_id = str(uuid.uuid4())
    disk_file_name = f'{temp_id}.{ext}'

    f.save(os.path.join(os.getenv('FILE_UPLOAD_STORAGE_PATH'), disk_file_name))

    file_in_db = File(
        id=temp_id,
        project_id=project_id,
        file_name_original=file_name,
        file_name_disk=disk_file_name,
        file_category=request.form.get('file_category', 'unspecified'),
        description=request.form.get('upload-description', ''),
    )
    db.session.add(file_in_db)
    db.session.commit()

    return redirect(url_for('projects.project_files_list', project_id=project_id))

@ProjectsBP.route('/createdummies')
def create_dummy_projects():
    proj = db.session.query(Project).filter_by(title=list(dummy_data.items())[0][0]).first()

    if not proj:
        for title, description in dummy_data.items():
            proj = Project(title = title, description = description, owner_id=session['user_id'])

            db.session.add(proj)
            db.session.commit()
    
    return '', 204

@ProjectsBP.route('/get')
def get_projects():
    proj = db.session.query(Project).all()

    out = {}
    for p in proj:
        out[p.title] = {'description': p.description, 'id': p.id}
    
    return jsonify(out), 200


# debug endpoint
@ProjectsBP.route('/getfiles/<project_id>')
def get_project_files_debug(project_id):
    files = db.session.query(File).filter(File.project_id == project_id).all()

    file_names = [{'file_name_original': f.file_name_original,
                   'file_name_disk': f.file_name_disk,
                   'file_category': f.file_category,
                   'description': f.description, 
                   'date': f.upload_date.date()} for f in files]
    
    return jsonify(file_names), 200

@ProjectsBP.route('/<project_id>/people/import', methods=['GET', 'POST'])
def people_import(project_id):
    if "user_id" not in session:
        return redirect(url_for("authentication.login", next=request.path))

    if not _user_has_project_access(session["user_id"], project_id):
        return render_template("error/404.html"), 404

    project = db.session.get(Project, project_id)
    if not project:
        return render_template("error/404.html"), 404

    if request.method == "GET":
        return render_template(
            "pages/people_import_upload.html",
            active_project_id=project_id,
            project=project,
        ), 200

    f = request.files.get("uploaded_file")
    if not f or not f.filename:
        return render_template(
            "pages/people_import_upload.html",
            active_project_id=project_id,
            project=project,
            error="Please choose a file.",
        ), 400

    ext = f.filename.split(".")[-1].lower()
    if ext not in ["csv", "xlsx"]:
        return render_template(
            "pages/people_import_upload.html",
            active_project_id=project_id,
            project=project,
            error="Only .csv or .xlsx is supported.",
        ), 400

    instance_dir = os.path.join(os.getcwd(), "instance")
    imports_dir = os.path.join(instance_dir, "imports")
    os.makedirs(imports_dir, exist_ok=True)

    temp_id = str(uuid.uuid4())
    temp_name = f"{temp_id}.{ext}"
    temp_path = os.path.join(imports_dir, temp_name)

    f.save(temp_path)

    if ext == "csv":
        headers, preview = _parse_csv_headers_preview(temp_path, preview_rows=10)
    else:
        headers, preview = _parse_xlsx_headers_preview(temp_path, preview_rows=10)

    if not headers:
        return render_template(
            "pages/people_import_upload.html",
            active_project_id=project_id,
            project=project,
            error="Could not read headers from file.",
        ), 400

    return render_template(
        "pages/people_import_map.html",
        active_project_id=project_id,
        project=project,
        temp_path=temp_path,
        headers=headers,
        preview_rows=preview,
    ), 200


@ProjectsBP.route('/<project_id>/people/import/commit', methods=['POST'])
def people_import_commit(project_id):
    if "user_id" not in session:
        return redirect(url_for("authentication.login", next=request.path))

    if not _user_has_project_access(session["user_id"], project_id):
        return render_template("error/404.html"), 404

    project = db.session.get(Project, project_id)
    if not project:
        return render_template("error/404.html"), 404

    temp_path = request.form.get("temp_path", "")
    if not temp_path or not os.path.exists(temp_path):
        return render_template("error/404.html"), 404

    ext = temp_path.split(".")[-1].lower()
    if ext == "csv":
        headers, rows = _read_all_csv_rows(temp_path)
    else:
        headers, rows = _read_all_xlsx_rows(temp_path)

    header_to_idx = {h: i for i, h in enumerate(headers)}

    def _idx_for(field_name: str):
        selected = request.form.get(field_name, "")
        if not selected or selected == "__none__":
            return None
        return header_to_idx.get(selected)

    name_i = _idx_for("map_name")
    email_i = _idx_for("map_email")
    phone_i = _idx_for("map_phone")
    title_i = _idx_for("map_title")
    role_level_i = _idx_for("map_role_level")

    def _get(row: list[str], idx: int | None):
        if idx is None:
            return None
        if idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        s = str(v).strip()
        return s if s != "" else None

    created_people = 0
    assigned = 0
    skipped = 0

    for row in rows:
        name = _get(row, name_i)
        email = _get(row, email_i)
        phone = _get(row, phone_i)
        title = _get(row, title_i)
        role_level = _get(row, role_level_i)

        if email:
            email = email.strip().lower()

        if not name and not email:
            skipped += 1
            continue

        if not name:
            name = "Unknown"

        person = None

        if email:
            person = db.session.execute(
                select(Person).where(Person.email == email)
            ).scalar_one_or_none()

            if not person:
                person = Person(name=name, email=email, phone=phone, title=title)
                db.session.add(person)
                db.session.flush()
                created_people += 1
        else:
            person = Person(name=name, email=None, phone=phone, title=title)
            db.session.add(person)
            db.session.flush()
            created_people += 1

        already_assigned = db.session.execute(
            select(
                exists().where(
                    and_(
                        ProjectPerson.project_id == project_id,
                        ProjectPerson.person_id == person.id,
                    )
                )
            )
        ).scalar()

        if not already_assigned:
            db.session.add(ProjectPerson(project_id=project_id, person_id=person.id, role_level=role_level))
            assigned += 1


    db.session.commit()

    try:
        os.remove(temp_path)
    except OSError:
        pass

    return redirect(url_for("dashboard.get_dashboard_project_people", project_id=project_id))

@ProjectsBP.route('/<project_id>/settings', methods=['GET'])
def project_settings(project_id):
    project = db.session.get(Project, project_id)
    if not project:
        return render_template("error/404.html"), 404

    return render_template(
        "pages/project_settings.html",
        active_project=project,
        active_project_id=project.id,
    ), 200
