import os
import csv
from datetime import date, datetime

from openpyxl import load_workbook

def path_to_file_from_disk(file_name_in_disk):
    '''
    build path to file in disk from file name
    '''
    return os.path.join(os.getenv('FILE_UPLOAD_STORAGE_PATH'), file_name_in_disk)


def _cell_to_text(value) -> str:
    '''
    Apparently the way it reads data auto converts dates into a datetime object,
    this function is just so our row strings that have dates only display year-month-day
    without time
    '''
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()

def parse_csv_headers_preview(filepath: str, preview_rows: int = 10):
    with open(filepath, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = []
        for _, row in zip(range(preview_rows), reader):
            rows.append(row)
        return headers, rows

def parse_xlsx_headers_preview(filepath: str, preview_rows: int = 10):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    headers = [_cell_to_text(h) for h in (headers_raw or [])]

    rows = []
    for _, r in zip(range(preview_rows), rows_iter):
        rows.append([_cell_to_text(v) for v in r])

    return headers, rows

def read_all_csv_rows(filepath: str):
    with open(filepath, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = [row for row in reader]
        return headers, rows

def read_all_xlsx_rows(filepath: str):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    headers = [_cell_to_text(h) for h in (headers_raw or [])]

    rows = []
    for r in rows_iter:
        rows.append([_cell_to_text(v) for v in r])

    return headers, rows
